from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
sheet = wb.active
sheet.title = "Economic Model"

# Header
sheet['A1'] = "Nébuleuse Bijoux — AI Agent Optimization Financial Model"
sheet['A1'].font = Font(bold=True, size=14)
sheet.merge_cells('A1:E1')

# Section 1: Current State
sheet['A3'] = "CURRENT STATE (Baseline)"
sheet['A3'].font = Font(bold=True, size=12, color="FFFFFF")
sheet['A3'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
sheet.merge_cells('A3:E3')

sheet['A4'] = "Metric"
sheet['B4'] = "Value"
sheet['C4'] = "Unit"
sheet['D4'] = "Source/Notes"

for cell in ['A4', 'B4', 'C4', 'D4']:
    sheet[cell].font = Font(bold=True, color="FFFFFF")
    sheet[cell].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# Current state data (use row references only for input)
sheet['A5'] = "AI Tickets Resolved/Month"
sheet['B5'] = 488
sheet['C5'] = "tickets"
sheet['D5'] = "Audit baseline (60 days)"

sheet['A6'] = "Automation Rate"
sheet['B6'] = '=B5/6029'
sheet['C6'] = "%"
sheet['D6'] = "488 / 6029 total tickets"

sheet['A7'] = "Cost per Ticket (One Pilot)"
sheet['B7'] = 2.05
sheet['C7'] = "€/ticket"
sheet['D7'] = "One Pilot contract rate"

sheet['A8'] = "Outsourcing Savings (Current)"
sheet['B8'] = '=B5*B7'
sheet['C8'] = "€/month"
sheet['D8'] = "488 tickets × €2.05"

sheet['A9'] = "AI Agent Subscription Cost"
sheet['B9'] = 589
sheet['C9'] = "€/month"
sheet['D9'] = "Current Gorgias plan"

sheet['A10'] = "Net Monthly Benefit (Current)"
sheet['B10'] = '=B8-B9'
sheet['C10'] = "€/month"
sheet['D10'] = "Savings - AI cost"

sheet['A11'] = "Annual Benefit (Current)"
sheet['B11'] = '=B10*12'
sheet['C11'] = "€/year"
sheet['D11'] = "Monthly × 12 months"

# Section 2: Target State
sheet['A14'] = "TARGET STATE (50% Automation)"
sheet['A14'].font = Font(bold=True, size=12, color="FFFFFF")
sheet['A14'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
sheet.merge_cells('A14:E14')

sheet['A15'] = "Metric"
sheet['B15'] = "Value"
sheet['C15'] = "Unit"
sheet['D15'] = "Source/Notes"

for cell in ['A15', 'B15', 'C15', 'D15']:
    sheet[cell].font = Font(bold=True, color="FFFFFF")
    sheet[cell].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

sheet['A16'] = "AI Tickets Resolved/Month"
sheet['B16'] = 1414
sheet['C16'] = "tickets"
sheet['D16'] = "Projection at 50%"

sheet['A17'] = "Automation Rate"
sheet['B17'] = '=B16/6029'
sheet['C17'] = "%"
sheet['D17'] = "1,414 / 6029 total tickets"

sheet['A18'] = "Cost per Ticket (One Pilot)"
sheet['B18'] = 2.05
sheet['C18'] = "€/ticket"
sheet['D18'] = "Same contract rate"

sheet['A19'] = "Outsourcing Savings (Target)"
sheet['B19'] = '=B16*B18'
sheet['C19'] = "€/month"
sheet['D19'] = "1,414 tickets × €2.05"

sheet['A20'] = "AI Agent Subscription Cost"
sheet['B20'] = 1035
sheet['C20'] = "€/month"
sheet['D20'] = "Target Gorgias plan tier"

sheet['A21'] = "Net Monthly Benefit (Target)"
sheet['B21'] = '=B19-B20'
sheet['C21'] = "€/month"
sheet['D21'] = "Savings - AI cost"

sheet['A22'] = "Annual Benefit (Target)"
sheet['B22'] = '=B21*12'
sheet['C22'] = "€/year"
sheet['D22'] = "Monthly × 12 months"

# Section 3: Summary Comparison
sheet['A25'] = "FINANCIAL IMPACT SUMMARY"
sheet['A25'].font = Font(bold=True, size=12, color="FFFFFF")
sheet['A25'].fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
sheet.merge_cells('A25:E25')

sheet['A26'] = "Metric"
sheet['B26'] = "Current"
sheet['C26'] = "Target"
sheet['D26'] = "Improvement"
sheet['E26'] = "% Change"

for cell in ['A26', 'B26', 'C26', 'D26', 'E26']:
    sheet[cell].font = Font(bold=True)
    sheet[cell].fill = PatternFill(start_color="FDD835", end_color="FDD835", fill_type="solid")

sheet['A27'] = "Monthly Outsourcing Savings"
sheet['B27'] = '=B8'
sheet['C27'] = '=B19'
sheet['D27'] = '=C27-B27'
sheet['E27'] = '=(C27-B27)/B27'

sheet['A28'] = "Monthly AI Agent Cost"
sheet['B28'] = '=B9'
sheet['C28'] = '=B20'
sheet['D28'] = '=C28-B28'
sheet['E28'] = '=(C28-B28)/B28'

sheet['A29'] = "Net Monthly Benefit"
sheet['B29'] = '=B10'
sheet['C29'] = '=B21'
sheet['D29'] = '=C29-B29'
sheet['E29'] = '=(C29-B29)/B29'

sheet['A30'] = "Annual Benefit"
sheet['B30'] = '=B11'
sheet['C30'] = '=B22'
sheet['D30'] = '=C30-B30'
sheet['E30'] = '=(C30-B30)/B30'

# Section 4: Full Outsourcing Context
sheet['A33'] = "FULL OUTSOURCING CONTEXT"
sheet['A33'].font = Font(bold=True, size=12, color="FFFFFF")
sheet['A33'].fill = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")
sheet.merge_cells('A33:E33')

sheet['A34'] = "Metric"
sheet['B34'] = "Current"
sheet['C34'] = "Target"
sheet['D34'] = "Notes"

for cell in ['A34', 'B34', 'C34', 'D34']:
    sheet[cell].font = Font(bold=True)
    sheet[cell].fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

sheet['A35'] = "Total One Pilot Monthly Spend"
sheet['B35'] = 12000
sheet['C35'] = 6000
sheet['D35'] = "€12k baseline, €6k at 50% automation"

sheet['A36'] = "Total AI Agent + One Pilot"
sheet['B36'] = '=B35+B9'
sheet['C36'] = '=C35+B20'
sheet['D36'] = "Combined monthly support cost"

sheet['A37'] = "Monthly Savings vs Baseline"
sheet['B37'] = '=B10'
sheet['C37'] = '=B35+B9-C35-B20'
sheet['D37'] = "Reduction in total support spend"

# Formatting
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in sheet.iter_rows(min_row=4, max_row=11, min_col=1, max_col=4):
    for cell in row:
        cell.border = thin_border
        if cell.column == 2 and row[0].row > 4:
            cell.number_format = '#,##0.00'

for row in sheet.iter_rows(min_row=15, max_row=22, min_col=1, max_col=4):
    for cell in row:
        cell.border = thin_border
        if cell.column == 2 and row[0].row > 15:
            cell.number_format = '#,##0.00'

for row in sheet.iter_rows(min_row=26, max_row=30, min_col=1, max_col=5):
    for cell in row:
        cell.border = thin_border
        if cell.column in [2, 3, 4]:
            cell.number_format = '#,##0.00'
        if cell.column == 5:
            cell.number_format = '0.0%'

for row in sheet.iter_rows(min_row=34, max_row=37, min_col=1, max_col=4):
    for cell in row:
        cell.border = thin_border
        if cell.column in [2, 3]:
            cell.number_format = '#,##0.00'

# Column widths
sheet.column_dimensions['A'].width = 38
sheet.column_dimensions['B'].width = 16
sheet.column_dimensions['C'].width = 16
sheet.column_dimensions['D'].width = 16
sheet.column_dimensions['E'].width = 16

wb.save('/sessions/gracious-elegant-hopper/mnt/Projects/Nébuleuse Bijoux/NB_Financial_Model.xlsx')
print("Excel file created successfully")
